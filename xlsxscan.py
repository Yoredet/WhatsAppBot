import openpyxl as op
import re
import os
from pprint import pprint
from time import time as t

filename = 'Книга1.xlsx'

raspisanie = {}
if os.path.isfile(filename):
    wb = op.load_workbook(filename, data_only=True)
    sheet = wb.active
    max_rows = sheet.max_row
    for i in range(7, max_rows+1):
        time = sheet.cell(row=i, column=5).value
        phone_number = sheet.cell(row=i, column=20).value
        yes_no_flag = sheet.cell(row=i, column=4).value
        fio = str(sheet.cell(row=i, column=15).value)
        fio_doctor = re.sub(r'[^\w\s]+|[\d]+', r'', fio).strip().title()
        if not time or yes_no_flag == 'ДА' or not phone_number or phone_number[3] != '9' or len(phone_number) != 17:
            continue
        phone = '+7' + re.sub(r'\W', '', phone_number)[1:]
        in_dict = {}
        in_dict[time] = phone
        if fio_doctor not in raspisanie:
            raspisanie[fio_doctor] = in_dict
        else:
            raspisanie[fio_doctor][time] = phone
    flag_dubl = ''

    for fio, graphic in raspisanie.items():
        for time, phone in graphic.items():
            if phone == flag_dubl:
                del graphic[time]
            else:
                flag = phone
    pprint(raspisanie)
else:
    print('Файл таблицы не загружен')
