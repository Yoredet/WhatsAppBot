from dotenv import load_dotenv
import openpyxl as op
import os
from pprint import pprint
import pywhatkit
from random import randint
import re
from secret import message_list
import telebot
from time import sleep

load_dotenv()
# Токен вашего бота
TOKEN = os.getenv('TOKEN')
bot = telebot.TeleBot(TOKEN)
raspisanie = {}


# Получение картинки
@bot.message_handler(content_types=['photo'])
def get_photo(message):
    bot.reply_to(message, 'Понял, принял')
    photo_info = message.photo[-1]
    file_id = photo_info.file_id
    file_object = bot.get_file(file_id)
    file_bytes = bot.download_file(file_object.file_path)
    target_file_name = "test.jpg"
    with open(target_file_name, 'wb') as writer:
        writer.write(file_bytes)


# Получение документа
@bot.message_handler(content_types=['document'])
def handle_file(message):
    global file_name
    file_name = message.document.file_name
    file_info = bot.get_file(message.document.file_id)
    downloaded_file = bot.download_file(file_info.file_path)
    with open(file_name, 'wb') as new_file:
        new_file.write(downloaded_file)
    bot.reply_to(message, 'Табличку получил')
    bot.send_message(message.chat.id, 'Ок')
    bot.stop_polling()


bot.polling()


def parsing_file():
    '''
    Парсинг файла
    '''
    wb = op.load_workbook(file_name, data_only=True)
    sheet = wb.active
    max_rows = sheet.max_row
    for i in range(7, max_rows+1):
        time = sheet.cell(row=i, column=5).value
        phone_number = sheet.cell(row=i, column=20).value
        yes_no_flag = sheet.cell(row=i, column=4).value
        fio = str(sheet.cell(row=i, column=15).value)
        fio_doctor = (re.sub(r'[^\w\s]+|[\d]+', r'', fio).strip().
                      title().replace('h', 'н'))
        fio_doctor = ' '.join(fio_doctor.split()[:3])
        if not time or yes_no_flag == 'ДА' or not phone_number or \
                phone_number[3] != '9' or len(phone_number) != 17:
            continue
        phone = '+7' + re.sub(r'\W', '', phone_number)[1:]
        in_dict = {}
        in_dict[time] = phone
        if fio_doctor not in raspisanie:
            raspisanie[fio_doctor] = in_dict
        else:
            raspisanie[fio_doctor][time] = phone
    flag_dubl = ''

    for fio, graphic in raspisanie.items():
        flag_dubl = ''
        to_del_list = []
        for time, phone in graphic.items():
            if phone == flag_dubl:
                to_del_list.append(time)
            else:
                flag_dubl = phone
        for i in to_del_list:
            del graphic[i]


def send_message_inst():
    '''
    Передача данных боту для отправки
    '''
    for fio_doctor, graphic in raspisanie.items():
        for time, phone in graphic.items():
            sluchai = randint(1, 3)
            message = message_list[sluchai].format(time, fio_doctor)
            pywhatkit.sendwhatmsg_instantly(phone_no=phone,
                                            message=message,
                                            tab_close=True)
            sleep(5)


def main():
    parsing_file()
    pprint(raspisanie)
    send_message_inst()
    if os.path.isfile(file_name):
        os.remove(file_name)


main()
